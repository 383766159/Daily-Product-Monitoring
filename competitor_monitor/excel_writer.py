"""Excel 写入模块 - 复刻参考项目格式"""

import os
import logging
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

SHEET_NAME = "\u7ade\u54c1\u76d1\u63a7"  # 竞品监控
DATA_START_ROW = 5

METRIC_LABELS = [
    "\u5212\u7ebf\u4ef7",            # 划线价
    "\u9875\u9762\u4ef7",            # 页面价
    "\u6d3b\u52a8\uff08\u4e13\u4eab/\u4f18\u60e0\u5238/LD/BD\uff09",  # 活动
    "\u8bc4\u5206",                  # 评分
    "\u8bc4\u8bba\u6570",            # 评论数
    "\u6392\u540d\uff08\u5927\u7c7b/\u5c0f\u7c7b\uff09",  # 排名
    "\u53d8\u4f53\u6570\u91cf",      # 变体数量
    "\u5e93\u5b58",                  # 库存
    "\u8d2d\u7269\u8f66\u5356\u5bb6",  # 购物车卖家
    "\u5176\u4ed6",                  # 其他
]

HEADER_FILL = PatternFill(start_color="DCE6F2", end_color="DCE6F2", fill_type="solid")


def today_str() -> str:
    d = date.today()
    return f"{d.year}/{d.month:02d}/{d.day:02d}"


def strip_money(value: str):
    if not value or value == "/":
        return "/"
    import re
    nums = re.findall(r"[\d.,]+", str(value))
    if not nums:
        return value
    numeric = nums[0].replace(",", "")
    try:
        return float(numeric)
    except ValueError:
        return value


def snapshot_by_metric(snapshot: dict) -> list:
    if not snapshot or not snapshot.get("ok"):
        return ["/"] * len(METRIC_LABELS)

    strike = strip_money(snapshot.get("strike_price", ""))
    page = strip_money(snapshot.get("page_price", ""))
    promotions = snapshot.get("promotions", "").replace("\n", " ").strip() or "/"
    rating = snapshot.get("rating", "") or "/"
    review_count = snapshot.get("review_count", "") or "/"
    rank_text = snapshot.get("rank", "") or "/"
    variation = snapshot.get("variation_count", "") or "/"
    inventory = snapshot.get("inventory", "") or "/"
    seller = snapshot.get("buy_box_seller", "") or "/"
    other = snapshot.get("other", "") or "/"

    return [
        strike if strike != "" else "/",
        page if page != "" else "/",
        promotions,
        rating,
        review_count,
        rank_text,
        variation,
        inventory,
        seller,
        other,
    ]


def _find_date_row(ws, date_str: str, last_col: int) -> int:
    """在 DATA_START_ROW 之后查找指定日期行。
    只检查实际有值的行，避免创建空单元格。
    """
    # 从 DATA_START_ROW 开始，每次跳 11 行（1 日期行 + 10 指标行）
    row = DATA_START_ROW
    max_check = max(ws.max_row + 11, DATA_START_ROW + 500)
    while row <= max_check:
        val = ws.cell(row=row, column=1).value
        if val is None:
            break  # 已经到底了
        if str(val).strip() == date_str:
            return row
        row += 11  # 跳到下一个日期块
    return 0  # 未找到


def _find_next_date_row(ws, last_col: int) -> int:
    """找到下一个可用的日期行位置"""
    # 从 DATA_START_ROW 开始，每次跳 11 行
    row = DATA_START_ROW
    max_check = max(ws.max_row + 11, DATA_START_ROW + 500)
    while row <= max_check:
        val = ws.cell(row=row, column=1).value
        if val is None:
            return row
        row += 11
    return row


def write_snapshots(excel_path: str, ordered_asins: list[str], snapshots: list[dict]):
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    by_asin = {s["asin"]: s for s in snapshots}
    date_str = today_str()
    last_col = 1 + len(ordered_asins)

    # ---- 写表头（第2-3行）----
    ws["A2"] = "\u7ade\u54c1 ASIN"
    ws["A3"] = "\u54c1\u724c"
    for idx, asin in enumerate(ordered_asins):
        col = idx + 2
        ws.cell(row=2, column=col, value=asin)
        brand = by_asin.get(asin, {}).get("brand", "")
        ws.cell(row=3, column=col, value=brand)

    # ---- 找今天是否已有日期行 ----
    date_row = _find_date_row(ws, date_str, last_col)
    if date_row == 0:
        date_row = _find_next_date_row(ws, last_col)

    # ---- 写日期行（合并 + 蓝底）----
    ws.cell(row=date_row, column=1, value=date_str)
    if last_col > 1:
        try:
            ws.merge_cells(start_row=date_row, start_column=1, end_row=date_row, end_column=last_col)
        except Exception:
            pass
    for c in range(1, last_col + 1):
        cell = ws.cell(row=date_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- 写 10 项指标 ----
    for metric_idx, label in enumerate(METRIC_LABELS):
        row_num = date_row + 1 + metric_idx
        ws.cell(row=row_num, column=1, value=label)
        for asin_idx, asin in enumerate(ordered_asins):
            snapshot = by_asin.get(asin)
            values = snapshot_by_metric(snapshot)
            ws.cell(row=row_num, column=asin_idx + 2, value=values[metric_idx])

    # ---- 列宽 ----
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    wb.save(excel_path)
    logger.info(f"Excel saved: {excel_path} ({len(ordered_asins)} ASINs, date row {date_row})")
